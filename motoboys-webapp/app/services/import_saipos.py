import datetime as dt
import io
from typing import Tuple

from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.models import Import, Ride
from app.services.courier_match import compute_fee_type, match_courier_id, norm_text, saipos_pending_reason
from app.services.week_service import get_open_week_for_date, get_or_create_week_for_date


def _find_col_alias(headers: list[str], canonical: str, aliases: list[str]) -> int:
    normalized = {norm_text(h): i for i, h in enumerate(headers) if h is not None}
    for candidate in [canonical, *aliases]:
        idx = normalized.get(norm_text(candidate))
        if idx is not None:
            return idx
    raise KeyError(canonical)


def _resolve_saipos_cols(headers: list[str]) -> tuple[int, int, int, int, int | None]:
    required = {
        "Id do pedido no parceiro": ["ID pedido", "Pedido parceiro", "Id pedido parceiro", "ID do pedido"],
        "Data da venda": ["Data venda", "Data do pedido", "Data", "Data/Hora"],
        "Entregador": ["Motoboy", "Entregador(a)", "Entregador nome"],
        "Valor Entregador": ["Valor do entregador", "Taxa entregador", "Valor motoboy", "Valor taxa motoboy"],
    }
    missing: list[str] = []
    out: dict[str, int] = {}
    for canonical, aliases in required.items():
        try:
            out[canonical] = _find_col_alias(headers, canonical, aliases)
        except KeyError:
            missing.append(canonical)

    cancel_idx = None
    try:
        cancel_idx = _find_col_alias(
            headers,
            "Está cancelado",
            ["Cancelado", "Pedido cancelado", "Está cancelada", "Status cancelado"],
        )
    except KeyError:
        cancel_idx = None

    if missing:
        raise HTTPException(
            status_code=400,
            detail={
                "error": "MISSING_REQUIRED_COLUMNS",
                "source": "SAIPOS",
                "missing": missing,
                "headers_found": headers,
            },
        )

    return (
        out["Id do pedido no parceiro"],
        out["Data da venda"],
        out["Entregador"],
        out["Valor Entregador"],
        cancel_idx,
    )


def _commit_rides_best_effort(db: Session, rides: list[Ride]) -> int:
    if not rides:
        return 0
    try:
        db.add_all(rides)
        db.commit()
        return len(rides)
    except IntegrityError:
        db.rollback()
        db.expunge_all()
        inserted = 0
        for r in rides:
            try:
                db.add(r)
                db.commit()
                inserted += 1
            except IntegrityError:
                db.rollback()
                continue
        return inserted


def import_saipos(db: Session, file_bytes: bytes, filename: str, file_hash: str) -> Tuple[str, int, int, int, int, list[str]]:
    imp = Import(source="SAIPOS", filename=filename, file_hash=file_hash, status="DONE", meta={})
    db.add(imp)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        existing = db.query(Import).filter(Import.source == "SAIPOS", Import.file_hash == file_hash).first()
        return str(existing.id), 0, 0, 0, int((existing.meta or {}).get("redirected_closed_week") or 0), []
    db.refresh(imp)

    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    header_row = None
    for r in range(1, 15):
        vals = [ws.cell(row=r, column=c).value for c in range(1, min(40, ws.max_column) + 1)]
        if any(v and norm_text(str(v)) in {"ENTREGADOR", "MOTOBOY"} for v in vals):
            header_row = r
            break
    if header_row is None:
        header_row = 1

    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        headers.append(str(v).strip() if v is not None else "")

    idx_id, idx_dt, idx_courier, idx_val, idx_cancel = _resolve_saipos_cols(headers)

    inserted = 0
    pend_assign = 0
    redirected_closed_week = 0
    week_ids_touched: set[str] = set()

    batch: list[Ride] = []

    for r in range(header_row + 1, ws.max_row + 1):
        external_id = ws.cell(row=r, column=idx_id + 1).value
        order_dt = ws.cell(row=r, column=idx_dt + 1).value
        courier_raw = ws.cell(row=r, column=idx_courier + 1).value
        value_raw = ws.cell(row=r, column=idx_val + 1).value

        if order_dt is None or value_raw is None:
            continue

        if isinstance(order_dt, str):
            parsed = None
            for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M:%S"):
                try:
                    parsed = dt.datetime.strptime(order_dt.strip(), fmt)
                    break
                except Exception:
                    pass
            if parsed is None:
                continue
            order_dt = parsed

        try:
            value_f = float(str(value_raw).replace(".", "").replace(",", "."))
        except Exception:
            continue

        fee_type = compute_fee_type(value_f)
        order_date = order_dt.date()
        week = get_or_create_week_for_date(db, order_date)
        week_ids_touched.add(str(week.id))

        paid_in_week_id = None
        if week.status != "OPEN":
            payable_week = get_open_week_for_date(db, dt.date.today())
            paid_in_week_id = payable_week.id
            week_ids_touched.add(str(payable_week.id))
            redirected_closed_week += 1

        courier_name_raw = str(courier_raw) if courier_raw is not None else None
        pending_special = saipos_pending_reason(courier_name_raw)

        courier_id = None
        status = "PENDENTE_ATRIBUICAO"
        pending_reason = pending_special if pending_special is not None else "NOME_NAO_CADASTRADO"

        if pending_special is None:
            courier_id, miss_reason = match_courier_id(db, courier_name_raw)
            if courier_id:
                status = "OK"
                pending_reason = None
            else:
                pending_reason = miss_reason or "NOME_NAO_CADASTRADO"

        is_cancelled = None
        if idx_cancel is not None:
            v = ws.cell(row=r, column=idx_cancel + 1).value
            if isinstance(v, str):
                is_cancelled = v.strip().upper().startswith("S")
            elif v is not None:
                is_cancelled = bool(v)

        ride = Ride(
            source="SAIPOS",
            import_id=imp.id,
            external_id=str(external_id) if external_id is not None else None,
            source_row_number=None,
            signature_key=None,
            order_dt=order_dt,
            delivery_dt=None,
            order_date=order_date,
            week_id=week.id,
            courier_id=courier_id,
            courier_name_raw=courier_name_raw,
            courier_name_norm=norm_text(courier_name_raw) if courier_name_raw is not None else None,
            value_raw=value_f,
            fee_type=fee_type,
            is_cancelled=is_cancelled,
            status=status,
            pending_reason=pending_reason,
            paid_in_week_id=paid_in_week_id,
            meta={"row": r},
        )
        batch.append(ride)
        if status.startswith("PENDENTE"):
            pend_assign += 1

        if len(batch) >= 500:
            inserted += _commit_rides_best_effort(db, batch)
            batch = []

    if batch:
        inserted += _commit_rides_best_effort(db, batch)

    try:
        imp_db = db.query(Import).filter(Import.id == imp.id).first()
        if imp_db is not None:
            meta = dict(imp_db.meta or {})
            meta["redirected_closed_week"] = int(redirected_closed_week)
            meta["week_ids_touched"] = sorted(week_ids_touched)
            imp_db.meta = meta
            db.commit()
    except Exception:
        db.rollback()

    return str(imp.id), inserted, pend_assign, 0, redirected_closed_week, sorted(week_ids_touched)
